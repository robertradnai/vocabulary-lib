import csv
import json
import logging
from .models import Flashcard, WordList, WordCollection, LearningProgressEntry, LearningProgress
from typing import Dict, Tuple
from openpyxl import Workbook
import pickle
from typing import List

mpl_logger = logging.getLogger('matplotlib')
mpl_logger.setLevel(logging.WARNING)

# Columns in the imported/exported excel sheets (A -> 1, B -> 2, ...)
LANG1_COL = 1
LANG2_COL = 2
REMARKS_COL = 3
LEARNING_STATUS_COL = 4


class WordListDaoExcel:

    def __init__(self, workbook_path: str):

        workbook = _load_workbook_by_path(workbook_path)

        self.word_collection: WordCollection = _excel_wb_to_word_collection(
            workbook=workbook,
            lang1_col=LANG1_COL,
            lang2_col=LANG2_COL,
            remarks_col=REMARKS_COL,
            learning_status_col=LEARNING_STATUS_COL
        )

    def get_word_list_names(self) -> List[str]:
        return list(self.word_collection.word_lists.keys())

    def get_word_list(self, word_list_name: str) -> WordList:
        return self.word_collection.word_lists[word_list_name]

    def is_word_list(self, word_list_name: str) -> bool:
        return word_list_name in self.word_collection.word_lists.keys()


def load_wordlist_book(wb_path: str) -> WordCollection:
    """
    Load the wordlist_book dictionary from an Excel file.
    :param wb_path: Path of Excel file
    """
    workbook = _load_workbook_by_path(wb_path)

    wordlist_book = _excel_wb_to_word_collection(
        workbook=workbook,
        lang1_col=LANG1_COL,
        lang2_col=LANG2_COL,
        remarks_col=REMARKS_COL,
        learning_status_col=LEARNING_STATUS_COL
    )
    return wordlist_book

def save_wordlist_book(wb_path: str, word_collection: WordCollection):
    """Save the word collection.
    :param wb_path: Path of the Excel workbook to be created
    :param word_collection: Object to be saved as Excel workbook
    """
    columns = {"lang1": LANG1_COL, "lang2": LANG2_COL, "remarks": REMARKS_COL, "learning status": LEARNING_STATUS_COL}


    workbook_new = Workbook()

    if "Sheet" in workbook_new.sheetnames:
        workbook_new.remove(workbook_new["Sheet"])

    for sheet_name in word_collection.word_lists.keys():
        workbook_new.create_sheet(sheet_name)
        _word_collection_to_wb(workbook_new, word_collection.word_lists[sheet_name], columns)

    _save_workbook(wb_path, workbook_new)


def _load_workbook_by_path(pathname: str):
    # Loading dictionary from file
    from openpyxl import load_workbook  # TODO revise this
    return load_workbook(pathname)


def _save_workbook(wb_path: str, altered_workbook):
    altered_workbook.save(filename=wb_path)


def _excel_wb_to_word_collection(workbook, lang1_col, lang2_col,
                                 remarks_col, learning_status_col) -> WordCollection:
    wordlist_book = {}
    for sheet_name in workbook.sheetnames:
        wordlist_frame: WordList = _excel_worksheet_to_wordlist(workbook, sheet_name, lang1_col, lang2_col,
                                                      remarks_col, learning_status_col)
        if len(wordlist_frame.flashcards) >= 5:
            wordlist_book[sheet_name] = wordlist_frame
    if len(wordlist_book) == 0:
        raise NoValidWordListsError("The selected file doesn't contain any valid word lists.")
    return WordCollection(
        lang1="lang1_placeholder",
        lang2="lang2_placeholder",
        word_lists=wordlist_book
    )


def _excel_worksheet_to_wordlist(workbook, sheet_name, lang1_col, lang2_col, remarks_col,
                                 learning_status_col) -> WordList:
    """
    Create a WordList object from a worksheet
    """
    flashcards: Dict[int, Flashcard] = {}
    learning_progress_codes: Dict[int, int] = {}

    # Avoiding overly large word lists
    if workbook[sheet_name].max_row > 10000:
        raise ValueError("Error: number of rows > 10000")

    # First line is for language information
    lang1 = workbook[sheet_name].cell(row=1, column=lang1_col).value
    lang2 = workbook[sheet_name].cell(row=1, column=lang2_col).value

    # Getting information from all the rows one by one
    for row in range(2, workbook[sheet_name].max_row+1):
        # Loading the proper cells to the function
        lang1_word = workbook[sheet_name].cell(row=row, column=lang1_col).value
        lang2_word = workbook[sheet_name].cell(row=row, column=lang2_col).value
        remarks = workbook[sheet_name].cell(row=row, column=remarks_col).value
        # Using learning status if there's a number between 0 and 1 in that column
        learning_status = workbook[sheet_name].cell(row=row, column=learning_status_col).value

        # Checking the content of the selected cells
        # Language cells must be filled in
        if (lang1_word == "") or (lang2_word == "") or (lang1_word is None) or (lang2_word is None):
            continue  # Skipping line if one of them is empty

        if remarks is None:
            remarks = ""

        flashcards[row] = Flashcard(
            lang1=lang1_word,
            lang2=lang2_word,
            remarks=remarks
        )

        learning_progress_codes[row] = learning_status

    return WordList(lang1=lang1, lang2=lang2, name=sheet_name, flashcards=flashcards,
                    learning_progress_codes=learning_progress_codes)


def _word_collection_to_wb(workbook, word_list: WordList, columns):

    vocab_name = word_list.name
    for row, flashcard in word_list.flashcards.items():
        # Iterate through rows in a worksheet
        workbook[vocab_name].cell(row=row, column=LANG1_COL).value = flashcard.lang1
        workbook[vocab_name].cell(row=row, column=LANG2_COL).value = flashcard.lang2
        workbook[vocab_name].cell(row=row, column=REMARKS_COL).value = flashcard.remarks
        workbook[vocab_name].cell(row=row, column=LEARNING_STATUS_COL).value = word_list.learning_progress_codes[row]


def save_string(file_path, data):
    with open(file_path, mode='w+') as f:
        f.write(data)


def load_string(file_path):
    with open(file_path, mode='r') as f:
        data = f.read()
    return data


class NoValidWordListsError(Exception):
    pass


def word_collection_to_pickle(path: str, word_collection: WordCollection):

    with open(path, 'wb') as f:
        pickle.dump(word_collection, f)


def word_collection_from_pickle(path: str) -> WordCollection:
    import pickle
    with open(path, 'rb') as f:
        return pickle.load(f)


def build_word_list_csv(lang1, lang2, flashcards_csv_str, learning_progress_json_str= None):

    flashcards = read_flashcards_from_csv_string(flashcards_csv_str, True)

    if learning_progress_json_str is None:
        learning_progress_codes = None
    else:
        learning_progress_codes_raw = json.loads(learning_progress_json_str)
        learning_progress_codes = {int(k): int(v) for k, v in
                                   learning_progress_codes_raw.items()}

    return WordList(lang1=lang1, lang2=lang2, flashcards=flashcards,
                    learning_progress_codes=learning_progress_codes)


def save_word_list_learning_progress_json(learning_progress: Dict[int, int]) -> str:
    return json.dumps(learning_progress)


def read_flashcards_from_csv_string(csv_string: str, ignore_first_line):
    from io import StringIO
    f = StringIO(csv_string)

    csv_rows = list(csv.reader(f, delimiter=','))

    def row_to_flashcard(row):
        lang1 = row[0]
        lang2 = row[1]
        remarks = row[2] if len(row) > 2 else ""

        return Flashcard(
            lang1=lang1,
            lang2=lang2,
            remarks=remarks
        )

    return {row_index: row_to_flashcard(row) for row_index, row
            in enumerate(csv_rows)}
